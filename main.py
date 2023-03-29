''' INFORMAÇÕES GERAIS NO FINAL DO CÓDIGO '''
#packages
import os
import csv
import time
import random
import openpyxl
from datetime import datetime
import pandas as pd
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook

#--------------------------------------------FIRST PART -> PREPARING THE PAGE----------------------------------------------------------#

# for making the right search

base_url = 'https://www.linkedin.com/jobs/search'
params = {
    'keywords': 'Publicidade',
    'location': 'Brasil',
    'geoId': '106057199',
    'trk': 'public_jobs_jobs-search-bar_search-submit',
    'position': '1',
    'pageNum': '0'
}

response = requests.get(base_url, params=params)
url = response.url

# open chromedriver and direct it to the right page
driver = webdriver.Chrome(executable_path=r'chrome-webdriver\chromedriver.exe')
driver.implicitly_wait(10)
driver.get(url)

# getting the number of jobs
number_of_jobs = driver.find_elements(By.CSS_SELECTOR, '.results-context-header__job-count')[0].text
number_of_jobs = number_of_jobs.replace('.', '')
print(f'Número de empregos: {number_of_jobs}')

# preparing and scrolling the page
print(f'Tempo de espera aproximado: {int(int(number_of_jobs)/25)*3} segundos. ---> Se o tempo for maior que 160 segundos desconsidere essa inform.\n')

# that is because when the number of results is bigger than 1000 Linkein will block new results
stop = min(2, int(int(number_of_jobs) / 25) + 1)  # you can try it up to 39... than probably the Linkedin will block the actions
for i in range(stop):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    try:
        button = driver.find_element(By.XPATH, "//button[@aria-label='Ver mais vagas']")
        actions = ActionChains(driver)
        actions.move_to_element(button).click().perform()
        time.sleep(random.uniform(2.6, 3.8)) #simulates a user in a better way
    except:
        time.sleep(random.uniform(2.6, 3.8))
        pass

#---------------------------------------SECOND PART -> CREATING NECESSARY LISTS------------------------------------------------------------------------------#

# creating empty lists
hrefList = []
jobtitle= []
companyname= []
company_URL= []
jobmodel= []
jobworkload= []
jobexp= []
numberofapplicants= []
posted_at=[]
gettingdata_at= [] # date
#---------------------------- 
# This part will be done later!!!!!!!!!!!
companysize = []
companyfollowers= []
companyheadquarters= []
#-----------------------------

#---------------------------------------------THIRD PART -> CREATING CVS, XLSX, CHECKING LIST IF ITS EXIST--------------------------------------------#
# name to files
file_name = f"jobs_{params['keywords']}"

# Verificar se o arquivo já existe antes de criar o CSV
if not os.path.exists(f'{file_name}.csv'):
    with open(f'{file_name}.csv', mode='w', encoding='utf-8', newline='') as file:
        writer = csv.writer(file)
        header =['Link', 'Título do Cargo', 'Empresa', 'Carga Horária', 'Experiência Requerida', 'Número de Candidatos', 'Data de Publicação', 'Data de Coleta']

        writer.writerow(header)

# Verificar se o arquivo já existe antes de criar o XLSX
if not os.path.exists(f'{file_name}.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados de Emprego"
    header = header
    ws.append(header)
    wb.save(f'{file_name}.xlsx')

# Checking if the link was visited
visited_links = []
with open(f'{file_name}.csv', mode='r', encoding='utf-8', newline='') as file:
    reader = csv.reader(file)
    try:
        next(reader) # pula a primeira linha
    except StopIteration:
        print("Arquivo vazio!")
    for links in reader:
        visited_links.append(links[0]) # Add links
        print("Links visitados:", visited_links)

#-------------------------4TH -> ADD LINKS OF JOB PAGES, OPEN IT, APPEND ELEMENTS TO THE LISTS, CREATE DICT AND DATAFRAMES --------------------------------------------------#

#Find job links and append it to a hreflist list
jobList = driver.find_elements(By.CSS_SELECTOR, '.base-card__full-link')
for job in jobList:
    href = job.get_attribute('href')
    if href not in visited_links:
        hrefList.append(href)

# Open links
for href in hrefList:
    driver.execute_script(f"window.open('{href}', '_blank');")
    
    #getting elements
    try:
        title = driver.find_element(By.CSS_SELECTOR, '.top-card-layout__title').text
        company = driver.find_element(By.CSS_SELECTOR, '.topcard__org-name-link').text
        load = driver.find_element(By.CSS_SELECTOR, 'li.description__job-criteria-item:nth-of-type(1) span.description__job-criteria-text').text.strip()
        xp = driver.find_element(By.CSS_SELECTOR, 'li.description__job-criteria-item:nth-of-type(2) span.description__job-criteria-text').text.strip()
        applicants= driver.find_element(By.CSS_SELECTOR, '.num-applicants__caption').text
        posted= driver.find_element(By.CSS_SELECTOR, '.posted-time-ago__text').text
        gettingdata=  datetime.now()
        gettingdata_formatted = gettingdata.strftime("M %M H %H -- %d/%m/%Y")

        #cleaing  append it to the lists
        jobtitle.clear()
        jobtitle.append(title)
        companyname.clear() 
        companyname.append(company)
        jobworkload.clear()
        jobworkload.append(load)
        jobexp.clear()
        jobexp.append(xp)
        numberofapplicants.clear()
        numberofapplicants.append(applicants)
        posted_at.clear()
        posted_at.append(posted)
        gettingdata_at.clear()
        gettingdata_at.append(gettingdata_formatted)
        
        # Create a dictionary list to store the job data 
        job_data = []
        # Create a dictionary from the lists
        job_dict = {'Link': href, 
            'title': jobtitle, 
            'company': companyname, 
            'workload': jobworkload, 
            'experience': jobexp, 
            'applicants': numberofapplicants,
            'posted_at': posted_at,
            'gettingdata_at': gettingdata_at}

        # Append the dictionary to the list
        job_data.append(job_dict)

        #Create a DataFrame from the dictionary
        job_data_df = pd.DataFrame(job_dict)

        # checking if it have every element
        if job_data_df.isna().any().any():
            print('Falha pois existem colunas não preenchidas')
        else:

#-----------------------------------------------------5TH -> FINALLY SAVE TO CSV AND XLSX FILES --------------------------------------------#
    
        # append the data to CSV
            job_data_df.to_csv(f'{file_name}.csv', mode='a', header=False, index=False)

        # Adiciona as informações do DataFrame ao arquivo "jobs.xlsx"
            book = openpyxl.load_workbook(f'{file_name}.xlsx')
            writer = pd.ExcelWriter(f'{file_name}.xlsx', engine='openpyxl') 
            writer.book = book
            if 'Dados de Emprego' in book.sheetnames:
            # adiciona ao final da planilha
                job_data_df.to_excel(writer, index=False, header=False, startrow=len(book['Dados de Emprego']['A'])+1, sheet_name='Dados de Emprego')
            else:
            # error 
                print ('Não foi possível salvar o excel')
            writer.save()

        #add to visited and clean job_data and the dataframe
        visited_links.append(href)
        job_data.clear()
        job_dict.clear()
        job_data_df.drop(job_data_df.index, inplace=True)
                  
    except:
       print("fail")
       visited_links.append(href)
       job_data.clear()
       job_dict.clear()
       job_data_df.drop(job_data_df.index, inplace=True)    

    # Wait for the tab to load before closing it
    time.sleep(random.uniform(4.1, 4.7))

    # clean cookies Close the tab 
    
    driver.close()

    # Switch back to the first tab
    driver.switch_to.window(driver.window_handles[0])

driver.quit()
'''
Bibliotecas externas necessárias: openpyxl, pandas, requests, selenium
OBS: a pasta chrome-webdriver com chromedriver deve estar no mesmo diretório desse script python

O que esse script faz:
1 - forma um link para a página pública de pesquisa do linkedin de acordo com os parametros 
    que você colocar no começo. Será feita uma pesquisa com esses parâmetros na próxima etapa - a biblioteca requests 
    facilita a formação do link de maneira correta
2 - abre uma janela do chromedriver, coloca o link da pesquisa desejada
3 - Rola essa janela, e quando necessário clicar no botão de mais resultados. O número de rolagens é regulável.
4 - cria uma série de listas vazia que serão usadas
5 - checar se os arquivos que csv e xlsx existem,  se eles existem adiciona os links do csv a lista de links visitados. se não existem 
    esses arquivos são criados com seus nomes ajustados pelos pelo parâmetro de pesquisa keyword
6 - verifica se links da página já estão na lista de checados antes de fazer as interações
7 - abre os links válidos um de cada vez em uma nova guia nova que é fechada após as operações necessárias
8 - na guia aberta faz o scraping dos dados
9 - cria um dicionário, e popula  as listas
10 - cria um dataframe
11 - finalmente salva os dados nos arquivos csv e xlsx criados e limpa os dados para próxima interação
12 - repete o processo até todos os links desejados serem visitados


Possíveis problemas: 
1 - O linkedin limita o número de empregos mostrados nas pesquisas, ele atualiza 25 de cada vez até o número 999...
    portanto não será possível obter um número maior que esse de links
2 - Se o scraping for feito muitas vez provavelmente o site irá direcionar para uma página de auth, isso é uma política contra ações 
    mal intencionadas..
    possível solução é usar proxy e ips diferentes
3 - o desempenho não é o ideal... uma maneira melhor de se fazer isso é pegar as respostas geradas pela aba network que da pagina que
    pode ser visualizada no espaço de desenvolvedor.... a página recebe links ao ser rolada. Nesses links estão respostas em html.
    apenas o número final muda aumentando em 25 por vez nesses links.... seria uma forma mais rápida que o selenium para fazer a 
    primeira parte do trabalho.
    Isso também pode ser adaptado para a segunda parte.
4 - Dataframes com colunas vazias ainda estão sendo salvos nas tabelas apesar das verificações que coloquei, essa verificação deve ser mudada
    para um código plenamente funcional... no entanto ter esses links de empregos em mão pode ser positivo. Basicamente o problema ocorre
    pois o linkedin depois de um certo número de solicitação direciona para a pág de auth. Se a ideia é não gravar o link dessas solicitações
    o código deve ser modificado. Uma ideia seria colocar um if statement para verificar se existe alguma coluna vazia... a checagem a seguir
    não funcionou -------->
    if job_data_df.isna().any().any():
            print('Falha pois existem colunas não preenchidas')
        else:
'''